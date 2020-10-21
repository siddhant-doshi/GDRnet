pip install dgl-cu101

cd /content/drive/My Drive/Work_Drug_Repurpose

import torch
import torch.nn as nn
import torch.nn.functional as F
import dgl
from dgl.nn.pytorch import GraphConv
from dgl.nn.pytorch import SAGEConv
from dgl.nn.pytorch import GATConv
from dgl import DGLGraph
import networkx as nx
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import openpyxl
import re
import numpy as np
import random
import csv 
from collections import Counter
import pickle
import math
from scipy import sparse
import seaborn as sns
import statistics
from prettytable import PrettyTable
import sklearn.metrics
from collections import OrderedDict
import json
import subprocess
import sys
import time
import xml.etree.ElementTree

#----------------------------------ALL Functions------------------------------------
#-----------------------------------------------------------------------------------

def extract(elem, tag, drop_s):
  text = elem.find(tag).text
  if drop_s not in text: raise Exception(text)
  text = text.replace(drop_s, "")
  try:
    return int(text)
  except ValueError:
    return float(text)

def gpu_util():
  i = 0

  d = OrderedDict()
  d["time"] = time.time()

  cmd = ['nvidia-smi', '-q', '-x']
  cmd_out = subprocess.check_output(cmd)
  gpu = xml.etree.ElementTree.fromstring(cmd_out).find("gpu")

  util = gpu.find("utilization")
  d["gpu_util"] = extract(util, "gpu_util", "%")

  d["mem_used"] = extract(gpu.find("fb_memory_usage"), "used", "MiB")
  d["mem_used_per"] = d["mem_used"] * 100 / 11171

  if d["gpu_util"] < 15 and d["mem_used"] < 2816 :
	  msg = 'GPU status: Idle \n'
  else:
	  msg = 'GPU status: Busy \n'

  now = time.strftime("%c")
  print('\n\nUpdated at %s\n\nGPU utilization: %s %%\nVRAM used: %s %%\n\n%s\n\n' % (now, d["gpu_util"],d["mem_used_per"], msg))

def get_relation_types(entity1,entity2,database = 'all'): #string matching using regex
  #Get the relation types of specific entities
  #database parameter defined realtions specific to a database {Drugbank,Hetionet,GNBR etc}
  lst = []
  tsv_file = open(relations_list_file)
  relations_tsv = csv.reader(tsv_file, delimiter="\t")
  if (database == 'all'):
    for line in relations_tsv:
      lst.append(line[0]) if (re.search(r"({0}:{1}|{1}:{0})".format(entity1,entity2),line[0],re.I)) else 0
  else :
    for line in relations_tsv:
      if (re.search(r"{0}::[\s\w+->]+::?({1}:{2}|{2}:{1})$".format(database,entity1,entity2),line[0],re.I)):
        lst.append(line[0])
  tsv_file.close()
  return lst

def get_links(relation_types): #send the exact relation string in a list
  links = []
  tsv_file = open(triplets_file)
  triplets_tsv = csv.reader(tsv_file, delimiter="\t")
  for line in triplets_tsv:
    if (not (line[0]==line[2])): links.append((line[0],line[2],line[1])) if (line[1] in relation_types) else 0
  tsv_file.close()
  return links

def encode(lst,masking): #this just takes a iterable and assigns numbers to it as in ids based on the maskings
  x = dict ()
  if (not (masking is None)):
    arr = masking
    for elem in lst:
      for mask in masking:
        if (elem in mask):
          break
      else :
        arr.append([elem])
    for elem in lst:
      for i in range(len(arr)):
        if (elem in arr[i]):
          break
      x[elem] = i

  else :
    for i in range(len(lst)): 
      x[list(lst)[i]] = i
  return x

def replace_links_with_ids(links,nodes_mapping,relation_types_mapping): 
  replaced_links = [(nodes_mapping[link[0]],nodes_mapping[link[1]],relation_types_mapping[link[2]]) for link in links]
  return replaced_links

def construct_DGL_graph(num_nodes,links,testing_links,is_relational_graph):
  links = torch.tensor(links)
  print (links.shape,links[0,:])
  testing_links = torch.LongTensor(testing_links)
  print (testing_links.shape,testing_links[0,:])
  G = dgl.graph((links[:,0],links[:,1]))
  G.add_edges(links[:,1],links[:,0])
  G_test = dgl.DGLGraph()
  G_test.add_nodes(num_nodes)
  G_test.add_edges(testing_links[:,0],testing_links[:,1])
  G_test.add_edges(testing_links[:,1],testing_links[:,0])
  edge_type = []
  #removing the test links
  print ("Added all the edges....Now removing the to be tested")
  i=0
  for link in testing_links:
    while (G.has_edges_between(link[0],link[1])):
      G.remove_edges(G.edge_ids(link[0],link[1]))
      G.remove_edges(G.edge_ids(link[1],link[0]))
  return G,G_test,edge_type

def shuffle_edge_ids(old_dct,relations_required_for_learning):
  replacement_mapping = dict ()
  for i in range(len(relations_required_for_learning)):
    replacement_mapping[old_dct[relations_required_for_learning[i]]] = i
  z = list(replacement_mapping.items())
  for itr in z:
    replacement_mapping[itr[1]] = itr[0]
  for item in old_dct.items():
    old_dct[item[0]] = replacement_mapping[item[1]]
  return old_dct 

def add_covid_links(graph,nodes_mapping):
  wb = openpyxl.load_workbook("/content/drive/My Drive/Work_Drug_Repurpose/Covid_Data/Covid-129Links_6Dis_118HostGenes.xlsx")
  sheet = wb.active
  links = []
  for i in range(1,sheet.max_row):
    links.append((sheet.cell(row=i,column=1).value,sheet.cell(row=i,column=2).value))
  wb = openpyxl.load_workbook("/content/drive/My Drive/Work_Drug_Repurpose/Covid_Data/Covid-Links_27_ViralDisProteins_332_humanGenes.xlsx")
  sheet = wb.active
  for i in range(1,sheet.max_row):
    links.append((sheet.cell(row=i,column=1).value,sheet.cell(row=i,column=2).value))
  print ("Len of covid links == %d"%(len(links)))
  covid_nodes = set ()
  for link in links:
    covid_nodes.update([link[0],link[1]])
  print ("Len of covid nodes == %d"%(len(covid_nodes)))
  nodes_to_be_added = covid_nodes.difference(set(nodes_mapping.keys()))
  print ("nodes to be added == %d"%(len(nodes_to_be_added)))
  graph.add_nodes(len(nodes_to_be_added))
  last_value = sorted(nodes_mapping.values())[len(nodes_mapping)-1]
  for node in nodes_to_be_added:
    if (not (node in nodes_mapping.keys())):
      last_value+=1
      nodes_mapping[node] = last_value
  replaced_links = [(nodes_mapping[link[0]],nodes_mapping[link[1]]) for link in links]
  for link in replaced_links:
    if (not (graph.has_edge_between(link[0],link[1]))):
      graph.add_edge(link[0],link[1])
      graph.add_edge(link[1],link[0])
  return graph,nodes_mapping

def remove_omim_diseases(links):
  new_links = []
  for link in links:
    if (not ((re.search(r"omim",link[0],re.I)) or (re.search(r"omim",link[1],re.I)))) : new_links.append(link)
  return new_links

def add_favipiravir_umifenovir_camostat_links(links):
  links.append(('Compound::DB12466','Gene::316','Hetionet::CbG::Compound:Gene'))
  links.append(('Compound::DB12466','Gene::7498','Hetionet::CbG::Compound:Gene'))
  links.append(('Compound::DB12466','Gene::1558','Hetionet::CbG::Compound:Gene'))
  links.append(('Compound::DB12466','Gene::1571','Hetionet::CbG::Compound:Gene'))
  links.append(('Compound::DB12466','Gene::5243','Hetionet::CbG::Compound:Gene'))
  links.append(('Compound::DB12466','Gene::9356','Hetionet::CbG::Compound:Gene'))
  links.append(('Compound::DB12466','Gene::9376','Hetionet::CbG::Compound:Gene'))
  links.append(('Compound::DB12466','Gene::116085','Hetionet::CbG::Compound:Gene'))
  links.append(('Compound::DB13729','Gene::5644','Hetionet::CbG::Compound:Gene'))
  links.append(('Compound::DB13609','Gene::1576','Hetionet::CbG::Compound:Gene'))
  links.append(('Compound::DB13609','Gene::2328','Hetionet::CbG::Compound:Gene'))
  links.append(('Compound::DB13609','Gene::2326','Hetionet::CbG::Compound:Gene'))
  links.append(('Compound::DB13609','Gene::1571','Hetionet::CbG::Compound:Gene'))
  links.append(('Compound::DB13609','Gene::1544','Hetionet::CbG::Compound:Gene'))
  links.append(('Compound::DB13609','Gene::1565','Hetionet::CbG::Compound:Gene'))
  links.append(('Compound::DB13609','Gene::1559','Hetionet::CbG::Compound:Gene'))
  links.append(('Compound::DB13609','Gene::1577','Hetionet::CbG::Compound:Gene'))
  links.append(('Compound::DB13609','Gene::7364','Hetionet::CbG::Compound:Gene'))
  links.append(('Compound::DB13609','Gene::54600','Hetionet::CbG::Compound:Gene'))
  links.append(('Compound::DB14066','Gene::1576','Hetionet::CbG::Compound:Gene'))
  links.append(('Compound::DB14066','Gene::5243','Hetionet::CbG::Compound:Gene'))
  return links

def load_data(relations_required_for_graph,relations_required_for_learning,relation_masking = None,is_relational_graph=False):
  #relation_masking is a list of lists i.e inner list gives relations types that are to be assigned same ids
  relation_types_mapping = encode(relations_required_for_graph,relation_masking) #relation_types = dict
  if (is_relational_graph) : relation_types_mapping = shuffle_edge_ids(relation_types_mapping,relations_required_for_learning)
  links = get_links(relations_required_for_graph) #27,47,520 #possibility of duplicates
  print ("acquired links...")
  links = remove_omim_diseases(links)
  links = add_favipiravir_umifenovir_camostat_links(links)
  print ("added favipiravir-uminofevir-camostat links...")
  nodes = set ()
  for link in links: #getting all the nodes from the links
    nodes.update([link[0],link[1]])
  nodes_mapping = encode(nodes,None) #encoded nodes = dict
  links = replace_links_with_ids(links,nodes_mapping,relation_types_mapping)
  #get training links and labels
  learning_links,learning_labels = set(),[]
  reqd_learning_relation = [relation_types_mapping[rel] for rel in relations_required_for_learning]
  for link in links:
    if (link[2] in reqd_learning_relation):
      learning_links.add((link[0],link[1])) #using a set to remove the duplicate links
  learning_links =list(learning_links)
  print ("lenof learing links = %d"%(len(learning_links)))
  learning_labels = [1 for _ in range(len(learning_links))] #appending 1-->indicating label 1 --> presence of connection
  dim = round(0.9*len(learning_links))
  print ("Dim of training links = %d"%(dim))
  training_links = learning_links[:dim]
  training_labels = learning_labels[:dim]
  testing_links = learning_links[dim:]
  testing_labels = learning_labels[dim:]
  print ("Dim of training links = %d"%(len(training_labels)))  
  print ("Dim of testing links = %d"%(len(testing_labels)))
  
  #now construct a DGLGraph
  print ("Constructing Graph....")
  graph,graph_test,E = construct_DGL_graph(len(nodes_mapping),links,testing_links,is_relational_graph)
  return graph,graph_test,E,nodes_mapping,relation_types_mapping,training_links,training_labels,testing_links,testing_labels


def create_batches(graph,graph_test,training_links,training_labels,testing_links,testing_labels,batch_size,zero_labelled_samples):
  nodes_dis = set() 
  learning_links = torch.LongTensor(training_links+testing_links)
  learning_labels = torch.LongTensor(training_labels+testing_labels)
  for link in learning_links:
    nodes_dis.add(link[1])
  nodes_comp = all_drugs
  nodes_comp = torch.LongTensor(list(nodes_comp))
  nodes_dis = torch.LongTensor(list(nodes_dis))
  no_conn_links,lab = torch.empty(zero_labelled_samples,2).type(torch.LongTensor),torch.zeros(zero_labelled_samples).type(torch.LongTensor)
  count=0
  while (count<zero_labelled_samples):
    indx1 = torch.randint(0,len(nodes_comp),[1])
    indx2 = torch.randint(0,len(nodes_dis),[1])
    tnsr = torch.LongTensor([nodes_comp[indx1],nodes_dis[indx2]])
    if (not ((graph.has_edge_between(tnsr[0].item(),tnsr[1].item())) or (graph_test.has_edge_between(tnsr[0].item(),tnsr[1].item())))):
      no_conn_links[count] = tnsr
      count+=1
  t_size = len(training_links)
  dim = round(0.9*zero_labelled_samples)
  training_links = torch.cat((torch.LongTensor(training_links),no_conn_links[:dim]))
  training_labels = torch.cat((torch.LongTensor(training_labels),lab[:dim]))
  testing_links = torch.cat((torch.LongTensor(testing_links),no_conn_links[dim:]))
  testing_labels = torch.cat((torch.LongTensor(testing_labels),lab[dim:]))
  
  num_batches = len(training_links)//batch_size
  
  links_train1 = torch.LongTensor(training_links)
  labels_train1 = torch.LongTensor(training_labels)
  batches = torch.empty(num_batches,batch_size,2).type(torch.LongTensor)
  batch_labels = torch.empty(num_batches,batch_size).type(torch.LongTensor)
  pos_size = round(0.4*batch_size)
  neg_size = batch_size-pos_size
  for i in range(num_batches):
    indx1 = torch.randint(0,t_size,[pos_size])
    indx2 = torch.randint(t_size,len(links_train1),[neg_size])
    indx = torch.cat((indx1,indx2))
    tnsr = torch.empty(batch_size,2).type(torch.LongTensor)
    for j in range(batch_size):
      tnsr[j] = links_train1[indx[j]]
    batches[i] = tnsr
    tnsr = torch.empty(batch_size)
    for j in range(batch_size):
      tnsr[j] = labels_train1[indx[j]]
    batch_labels[i] = tnsr
    print (i)

  return batches,batch_labels,testing_links,testing_labels,training_links,training_labels

def get_binary_feature_matrix(total_nodes,feature_size):
  feat_mat = np.empty([total_nodes,feature_size],dtype=int)
  for i in range(total_nodes):
    b = bin(i)[2:].rjust(feature_size,'0')
    feat_mat[i] = list(map(int,[char for char in b]))
  return feat_mat

def save_variable(variable,filename):
  pickle.dump(variable,open(filename, "wb"))

def load_variable(filename):
  return pickle.load(open(filename,'rb'))

def get_node_name(id):
  return list(nodes_mapping.keys())[list(nodes_mapping.values()).index(id)]

def get_node_id(name):
  return list(nodes_mapping.values())[list(nodes_mapping.keys()).index(name)]

def get_approved_drugs():
  #getting list of approved drugs only
  sheet = openpyxl.load_workbook('/content/drive/My Drive/Work_Drug_Repurpose/Approved_drugs.xlsx').active
  approved_drugs = []
  for i in range(1,sheet.max_row+1):
    approved_drugs.append("Compound::"+sheet.cell(row=i,column=2).value)
  approved_drugs.append("Compound::DB01041")
  return approved_drugs

def get_disease_batches(nodes_mapping,disease_list): #disease_id in the form like Disease::MESH..
  dct = get_drug_name_desc_dict()
  keys = list(nodes_mapping.keys())
  drugs = []
  batches = []
  for key in keys:
    if (re.search(r"Compound+",key,re.I)):
      if (key in dct.keys()): 
        a = dct[key][1].split(',')
        if (not ((('experimental' in a) and (len(a)==1)) or 'withdrawn' in a)) :
          drugs.append(nodes_mapping[key])
  for disease in disease_list:
    disease_id = get_node_id(disease)
    batch = []
    for drug in drugs:
      batch.append((drug,disease_id))
    batches.append(batch)
  return batches,dct

def get_drkg_embeddings_feature_matrix(total_nodes,nodes_mapping):
  embeddings = np.load("/content/drive/My Drive/Work_Drug_Repurpose/DRKG_TransE_l2_entity.npy")
  drkg_encoding = dict ()
  tsv_file = open(entity_list_file)
  entities_tsv = csv.reader(tsv_file, delimiter="\t")
  for line in entities_tsv:
    drkg_encoding[line[0]] = int(line[1])
  mat = np.empty([total_nodes,embeddings.shape[1]])
  for i in range(total_nodes):
    print (i)
    mat[i] = embeddings[drkg_encoding[get_node_name(i)]]
  return mat

def get_drkg_relation_embedding(relation_name):
  embeddings = np.load("/content/drive/My Drive/Work_Drug_Repurpose/DRKG_TransE_l2_relation.npy")
  tsv_file = open(relations_list_file)
  relations_tsv = csv.reader(tsv_file, delimiter="\t")
  rel_embed = np.zeros(embeddings.shape[1])
  for line in relations_tsv:
    print (line[0])
    if (relation_name==line[0]):
      rel_embed = embeddings[int(line[1])]
      break
  return rel_embed

def get_drug_name_desc_dict():
  sheet = openpyxl.load_workbook('Drug_details.xlsx').active
  dct = dict ()
  for i in range(1,sheet.max_row+1):
    dct[sheet.cell(row=i,column=1).value] = (sheet.cell(row=i,column=2).value,sheet.cell(row=i,column=3).value,sheet.cell(row=i,column=4).value)
  return dct

def get_entity_list(entity,nodes_mapping):
  lst = []
  for key in nodes_mapping.keys():
    if (re.search(r"%s"%(entity),key,re.I)):
       lst.append(key)
  return lst

def save_model(the_model,path):
  torch.save(the_model.state_dict(),path)

def load_model(model,path):
  model.load_state_dict(torch.load(path))
  return model

def load_model_on_cpu(model,path):
  model.load_state_dict(torch.load(path,map_location=torch.device('cpu')))
  return model

def get_graph_stats(graph,nodes_mapping):
  dct = dict()
  dct['num_nodes'] = graph.number_of_nodes()
  dct['num_edges'] = graph.number_of_edges()
  dct['num_drugs'] = 0
  dct['num_diseases'] = 0
  dct['num_genes'] = 0
  dct['num_anatomies'] = 0
  for key in nodes_mapping.keys():
    if (re.search(r"Compound",key,re.I)) : dct['num_drugs']+=1
    elif (re.search(r"Disease",key,re.I)) : dct['num_diseases']+=1
    elif (re.search(r"Anatomy",key,re.I)) : dct['num_anatomies']+=1
    else : dct['num_genes']+=1
  return dct

def histogram(lst):
  n, bins, patches = plt.hist(x=lst, bins=1180, color='black')

def get_A_tilda(graph):
  adj = sparse.csr_matrix(graph.adjacency_matrix_scipy(),dtype=float)
  edges = list(graph.edges())
  for i in range(len(edges[0])):
    adj[edges[0][i].item(),edges[1][i].item()] = (graph.in_degree(edges[0][i].item())*graph.in_degree(edges[1][i].item()))**(-0.5)
  return adj

def get_rank_prob(lst,name):
  cnt = 0
  tup = (9000,0.0)
  for i in range(len(lst)):
    if(lst[i][0]==name):
      tup = (i+1,lst[i][1])
      break
  return tup

def get_ranks_covid_drugs(dct,probs = False):
  new_dct = dict()
  lst = sorted(dct.items(),key=lambda t:t[1])
  lst = lst[::-1] if (probs) else lst
  wb = openpyxl.load_workbook("Covid_clinical_drugs.xlsx").active
  for i in range(1,wb.max_row+1):
    new_dct[wb.cell(row=i,column=1).value] = get_rank_prob(lst,wb.cell(row=i,column=1).value)
  return new_dct

def excel_results(dictionaries,name):
  wb = openpyxl.Workbook()
  sheet = wb.active
  keys = list(dictionaries[0].keys())
  for i in range(1,len(dictionaries[0])+1):
    sheet.cell(row=i,column=1).value = keys[i-1]
    for j in range(2,len(dictionaries)+2):
      sheet.cell(row=i,column=j).value = dictionaries[j-2][keys[i-1]]
  wb.save(model_directory+name)

def excel_covid_clinical_results(dictionaries,p,name):
  wb = openpyxl.Workbook()
  sheet = wb.active
  covid_clinical_dict = [get_ranks_covid_drugs(dct,probs=p) for dct in dictionaries]
  keys = list(covid_clinical_dict[0].keys())
  for i in range(1,len(covid_clinical_dict[0])+1):
    sheet.cell(row=i,column=1).value = keys[i-1]
    k=2
    for j in range(len(covid_clinical_dict)):
      sheet.cell(row=i,column=k).value = covid_clinical_dict[j][keys[i-1]][0]
      k+=1
      sheet.cell(row=i,column=k).value = covid_clinical_dict[j][keys[i-1]][1]
      k+=1
  wb.save(model_directory+name)

def get_small_molecules(): #I am considering 8079 drugs out of which 7491 has molecular embedding
  f = open("/content/drive/My Drive/Work_Drug_Repurpose/Small_Molecules.txt","r")
  drugs = []
  for line in f:
    drugs.append("Compound::"+line[:7])
  return drugs

def get_entity_list(name,nodes_mapping,return_id=True):
  e = set()
  if (return_id):
    for key in nodes_mapping.keys():
      if(re.search(r"%s"%(name),key,re.I)) : e.add(nodes_mapping[key])
  else:
    for key in nodes_mapping.keys():
      if(re.search(r"%s"%(name),key,re.I)) : e.add(key)
  return e

def get_drug_map_inorder(nodes_mapping):
  small_molecules = get_small_molecules()
  mp,indx_to_be_taken = [],[]
  i=0
  for drug in small_molecules:
    if (drug in nodes_mapping.keys()):
      mp.append(nodes_mapping[drug])
      indx_to_be_taken.append(i)
    i+=1
  return torch.LongTensor(mp),indx_to_be_taken

def get_molecular_feat_combined(filename,nodes_mapping,X):
  feat = np.load(filename)
  mp,indx = get_drug_map_inorder(nodes_mapping)
  mol_feat_updated = torch.empty(len(indx),300)
  X_feat_updated = torch.empty(len(indx),400)
  for i in range(len(indx)):
    mol_feat_updated[i] = torch.tensor(feat[indx[i]])
    X_feat_updated[i] = X[mp[i]]
  return torch.cat((mol_feat_updated,X_feat_updated),dim=1)

def get_associated_genes(all_genes,entity_id,return_ids = True): #Give in a drug/disease entity id - if return_ids=False - we will get the names
  edges = list(graph_nx.edges(nbunch=entity_id))
  neighbours = set([edge[1] for edge in edges])
  genes_associated = neighbours.intersection(all_genes)
  return genes_associated

def remove_deg_zero_genes(gene_list):
  modified_list = []
  for gene in gene_list:
    if (gene_network.degree(gene)!=0) : modified_list.append(gene)
  return modified_list

def get_dctScore(drug_id,dis_id,all_genes,gene_network): #I will use the nx graph
  #Z_score will be calculated based on gene interactome of a drug and a disease
  dis_genes = remove_deg_zero_genes(get_associated_genes(all_genes,dis_id))
  drug_genes = remove_deg_zero_genes(get_associated_genes(all_genes,drug_id))
  if (len(dis_genes)==0 or len(drug_genes)==0):
    d_ct = 15000
  else:
    s = 0
    for c in dis_genes:
      a = []
      for t in drug_genes:
        try: #coz some genes have no path between them
          a.append(len(nx.shortest_path(gene_network,source=c,target=t))-1)
        except:
          a.append(20000)
      s+=min(a)
    for t in drug_genes:
      a = []
      for c in dis_genes:
        try:#coz some genes have no path between them
          a.append(len(nx.shortest_path(gene_network,source=c,target=t))-1)
        except:
          a.append(20000)
      s+=min(a)
    d_ct = s/(len(dis_genes)+len(drug_genes))
  
  return d_ct

def get_random_gene_list(degree_ref_list,degreewise_gene_dct): #degree distribution
  dist = Counter(degree_ref_list)
  lst = []
  for key in dist.keys():
    lst+=random.sample(degreewise_gene_dct[key],dist[key])
  return lst

def get_dct_from_genelists(l1,l2):
  s = 0
  for c in l1:
    a = []
    for t in l2:
      try:
        a.append(len(nx.shortest_path(gene_network,source=c,target=t))-1)
      except:
        a.append(20000)
    s+=min(a)
  for t in l2:
    a = []
    for c in l1:
      try:
        a.append(len(nx.shortest_path(gene_network,source=c,target=t))-1)
      except:
        a.append(20000)
    s+=min(a)
  d_ct = s/(len(l1)+len(l2))
  return d_ct

def get_Zscore(all_genes,dis_id,drug_id,degreewise_gene_dct,gene_network,num_seeds=50):
  C = remove_deg_zero_genes(get_associated_genes(all_genes,dis_id))
  T = remove_deg_zero_genes(get_associated_genes(all_genes,drug_id))
  d_ct_target = get_dctScore(drug_id,dis_id,all_genes,gene_network)
  print ("Drug_id(T) = %d | C = %d | T = %d | d_ct = %f"%(drug_id,len(C),len(T),d_ct_target))
  if (d_ct_target==15000):
    print ("Original d_ct unable to calculate")
    Z_score = 15000
  else:
    C_deg_list = [gene_network.degree(c) for c in C]
    T_deg_list = [gene_network.degree(t) for t in T]
    dct_rnd = []
    for i in range(num_seeds):
      C_rnd = get_random_gene_list(C_deg_list,degreewise_gene_dct)
      T_rnd = get_random_gene_list(T_deg_list,degreewise_gene_dct)
      dct_rnd.append(get_dct_from_genelists(C_rnd,T_rnd))
    #print (dct_rnd)
    try:
      Z_score = (d_ct_target-statistics.mean(dct_rnd))/(statistics.stdev(dct_rnd))
    except:
      Z_score = d_ct_target
  return Z_score

def get_degreewise_genedict(all_genes,gene_graph):
  dct = dict()
  for gene in all_genes:
    dct[gene_graph.degree(gene)] = []
  for gene in all_genes:
    dct[gene_graph.degree(gene)].append(gene)
  return dct

def standardize(t):
  mean = torch.mean(t)
  stdev = torch.std(t)
  standard_t = (t-mean)/stdev
  return standard_t

def get_subgraph(graph_nx,nodes): #give in a nx graph
  return graph_nx.subgraph(nodes).copy()

def atc_dict():
  wb = openpyxl.load_workbook("/content/drive/My Drive/Drug_ATC_updated.xlsx").active
  dct = dict()
  for i in range(1,wb.max_row+1):
    try:
      dct[wb.cell(row=i,column=1).value] = wb.cell(row=i,column=2).value.split(',')
    except AttributeError:
      dct[wb.cell(row=i,column=1).value] = ""
  return dct

def get_anatomies_for_genes(gene_id):
  anatomies = []
  for edge in graph_nx.edges(gene_id):
    if (edge[1] in all_anat): anatomies.append(edge[1])
  return set(anatomies)

def get_connected_drugs_for_disease(dis_id):
  drugs = []
  for edge in graph_nx.edges(dis_id):
    if (edge[1] in all_drugs): drugs.append(edge[1])
  return set(drugs) 

def get_common_anatomies(gene_list):
  if (len(gene_list)==1):
    return get_anatomies_for_genes(gene_list[0])
  elif (len(gene_list)==2):
    return get_anatomies_for_genes(gene_list[0]).intersection(get_anatomies_for_genes(gene_list[1]))
  else:
    f = get_anatomies_for_genes(gene_list[0]).intersection(get_anatomies_for_genes(gene_list[1]))
    for i in range(2,len(gene_list)):
      f = f.intersection(get_anatomies_for_genes(gene_list[i]))
    return f
  
def get_common_drugs(disease_list):
  if (len(disease_list)==1):
    return get_connected_drugs_for_disease(disease_list[0])
  elif (len(disease_list)==2):
    return get_connected_drugs_for_disease(disease_list[0]).intersection(get_connected_drugs_for_disease(disease_list[1]))
  else:
    f = get_connected_drugs_for_disease(disease_list[0]).intersection(get_connected_drugs_for_disease(disease_list[1]))
    for i in range(2,len(disease_list)):
      f = f.intersection(get_connected_drugs_for_disease(disease_list[i]))
    return f

def save_graph_data(path,graph,graph_test,nodes_mapping,relation_types_mapping,training_links,training_labels,testing_links,testing_labels,graph_nx):
  save_variable(graph,model_directory+"Data/graph.p")
  save_variable(graph_test,model_directory+"Data/graph_test.p")
  save_variable(nodes_mapping,model_directory+"Data/nodes_mapping.p")
  save_variable(relation_types_mapping,model_directory+"Data/relation_types_mapping.p")
  save_variable(training_links,model_directory+"Data/training_links.p")
  save_variable(training_labels,model_directory+"Data/training_labels.p")
  save_variable(testing_links,model_directory+"Data/testing_links.p")
  save_variable(testing_labels,model_directory+"Data/testing_labels.p")
  save_variable(graph_nx,model_directory+"Data/graph_nx.p")

def load_graph_data(model_directory):
  graph = load_variable(model_directory+"Data/graph.p")
  graph_test = load_variable(model_directory+"Data/graph_test.p")
  nodes_mapping = load_variable(model_directory+"Data/nodes_mapping.p")
  relation_types_mapping = load_variable(model_directory+"Data/relation_types_mapping.p")
  training_links = load_variable(model_directory+"Data/training_links.p")
  training_labels = load_variable(model_directory+"Data/training_labels.p")
  testing_links = load_variable(model_directory+"Data/testing_links.p")
  testing_labels = load_variable(model_directory+"Data/testing_labels.p")
  graph_nx = load_variable(model_directory+"Data/graph_nx.p")
  return graph,graph_test,nodes_mapping,relation_types_mapping,training_links,training_labels,testing_links,testing_labels,graph_nx

def get_sparse_tensor_from_scipy_coo(coo):
  values = coo.data
  indices = np.vstack((coo.row, coo.col))
  i = torch.LongTensor(indices)
  v = torch.FloatTensor(values)
  shape = coo.shape
  return torch.sparse.FloatTensor(i, v, torch.Size(shape))

def count_parameters(model):
  table = PrettyTable(["Modules", "Parameters"])
  total_params = 0
  for name, parameter in model.named_parameters():
    if not parameter.requires_grad: continue
    param = parameter.numel()
    table.add_row([name, param])
    total_params+=param
  print(table)
  print(f"Total Trainable Params: {total_params}")
  return total_params

#--------------------------------Declarations----------------------------------#
#--------- Get these 3 files from DRKG database, and give the path
relations_list_file = "relations.tsv"
entity_list_file = "entities.tsv"
triplets_file = "drkg.tsv"
#----------
covid_disease_file = "Covid_disease_list.xlsx"
wb = openpyxl.load_workbook(covid_disease_file).active
covid_disease_list = [wb.cell(row=i,column=1).value for i in range(1,wb.max_row+1)]
db_treats_relation_embedding = get_drkg_relation_embedding("DRUGBANK::treats::Compound:Disease")
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(device)

#get relation types
r1 = get_relation_types("Compound","Disease",database='Hetionet') #palliation and treatment
r2 = get_relation_types("Compound","Disease",database='Drugbank') #treatment
r3 = get_relation_types("Compound","Gene",database='Hetionet')
r4 = get_relation_types("Compound","Gene",database='Drugbank')
r6 = get_relation_types("Gene","Gene",database='Hetionet')
r8 = get_relation_types("Disease","Gene")
r9 = get_relation_types("Disease","Disease",database='Hetionet')
r10 = get_relation_types("Compound",'Compound',database='Hetionet')
r11 = get_relation_types("Disease","Anatomy")
r12 = get_relation_types("Gene","Anatomy")

req_links = r1+r3+r6+r8+r9+r10+r2+r4+r11+r12
learning_relations = r1+r2

graph,graph_test,E,nodes_mapping,relation_types_mapping,training_links,training_labels,testing_links,testing_labels = load_data(req_links,learning_relations)

get_graph_stats(graph,nodes_mapping)

graph_nx = graph.to_networkx().to_undirected()

#saving the data -- 
model_directory = 'SIGN_Models/Model_7/'
save_graph_data(model_directory,graph,graph_test,nodes_mapping,relation_types_mapping,training_links,training_labels,testing_links,testing_labels,graph_nx)

#data loading-- 
model_directory = 'SIGN_Models/Model_7/'
graph,graph_test,nodes_mapping,relation_types_mapping,training_links,training_labels,testing_links,testing_labels,graph_nx = load_graph_data(model_directory)

all_genes = get_entity_list("Gene",nodes_mapping) #by default we will get a set of gene_ids
all_drugs = get_entity_list("Compound",nodes_mapping)
all_dis = get_entity_list("Disease",nodes_mapping)
all_anat = get_entity_list("Anatomy",nodes_mapping)
gene_network = get_subgraph(graph_nx,list(all_genes))
degreewise_gene_dct = get_degreewise_genedict(all_genes,gene_network)
atc_dictionary = atc_dict()
drug_name_dict = get_drug_name_desc_dict()
covid_genes = set()
for dis in covid_disease_list[:27]:
  covid_genes.update(get_associated_genes(all_genes,get_node_id(dis)))
print (len(covid_genes))

batch_size = 512
zero_labelled_samples = 200000
batches,batch_labels,testing_links,testing_labels,training_links,training_labels = create_batches(graph,graph_test,training_links,training_labels,testing_links,testing_labels,batch_size,zero_labelled_samples)
num_batches = len(batches)

input_features = torch.FloatTensor(get_drkg_embeddings_feature_matrix(graph.number_of_nodes(),nodes_mapping))
save_variable(input_features,model_directory+"Data/input_features.p")

input_features = load_variable(model_directory+"Data/input_features.p")

#A_tilda = get_A_tilda(graph)
#save_variable(A_tilda,model_directory+"Data/A_tilda.p")
model_directory = 'SIGN_Models/Model_7/'
A_tilda = load_variable(model_directory+"Data/A_tilda.p")
ax = A_tilda*np.array(input_features)
a2x = A_tilda*ax
a2x = torch.tensor(a2x,dtype=torch.float)
ax = torch.tensor(ax,dtype=torch.float)

save_variable(ax,model_directory+"Data/ax.p")
save_variable(a2x,model_directory+"Data/a2x.p")

L_Relu = nn.LeakyReLU(0.2)
sig = nn.Sigmoid()
Relu = nn.ReLU()
tanh = nn.Tanh()

class SIGN(nn.Module):
  def __init__(self):
    super(SIGN, self).__init__()
    decoder_dim = 250
    input_dim = 400
    r = 3
    self.theta0 = nn.Linear(input_dim,decoder_dim) 
    self.theta1 = nn.Linear(input_dim,decoder_dim)
    self.theta2 = nn.Linear(input_dim,decoder_dim)
    self.combine1 = nn.Linear(decoder_dim*r,decoder_dim)
    self.layer8 = nn.Linear(decoder_dim,decoder_dim)
    self.layer9 = nn.Linear(decoder_dim,decoder_dim)
  
  def decoder(self,t,batch):
    self.t_new = torch.empty(len(batch)).to(device)
    for i in range(len(batch)):
      self.c = torch.dot(t[batch[i,0].item()],self.layer8(t[batch[i,1].item()])).to(device) #+torch.dot(t[batch[i,1].item()],self.layer9(t[batch[i,0].item()]))).to(device)
      self.t_new[i] = self.c
    return self.t_new

  def forward(self,X,ax,a2x,batch):
    t1 = tanh(self.theta0(X))
    t2 = tanh(self.theta1(ax))
    t3 = tanh(self.theta2(a2x))
    c = torch.cat((t1,t2,t3),dim=1)
    c = L_Relu(self.combine1(c))
    t1 = self.decoder(c,batch)
    return c,t1

class Seq_GNN(nn.Module):
  def __init__(self):
    super(Seq_GNN, self).__init__()
    d = 400
    decoder_dim = 250
    self.layer1 = GATConv(d,320,1,allow_zero_in_degree=True)
    self.layer2 = GATConv(320,decoder_dim,1,allow_zero_in_degree=True)
    self.layer8 = nn.Linear(decoder_dim,decoder_dim)
	
  def decoder(self,t,batch):
    self.t_new = torch.empty(len(batch)).to(device)
    for i in range(len(batch)):
      self.c = torch.dot(t[batch[i,0].item()],self.layer8(t[batch[i,1].item()])).to(device) #+torch.dot(t[batch[i,1].item()],self.layer9(t[batch[i,0].item()]))).to(device)
      self.t_new[i] = self.c
    return self.t_new
  
  def forward(self,g,X,batch):
    c = L_Relu(self.layer1(g,X))
    c = L_Relu(self.layer2(g,c.squeeze()))
    t1 = self.decoder(c.squeeze(),batch)
    return c.squeeze(),t1

neg,pos=0,0
for labels in batch_labels:
  dct = Counter(np.array(labels))
  neg+=dct[0]
  pos+=dct[1]
print (neg/num_batches,pos/num_batches)
w = neg/pos
print (w)

net = SIGN().to(device)
#net = Seq_GNN().to(device)
optimizer = torch.optim.Adam(net.parameters(), lr=0.0015)
BCEloss = nn.BCEWithLogitsLoss(pos_weight = torch.Tensor([w]).to(device))
epoch_loss = []
training_acc = []
avg_epoch_loss = []
F1_scores = []
count_parameters(net)

#training ----------------------------------
net.train()
num_epochs = 1

for epoch in range(num_epochs):
  batch_loss = []
  test_loss = []
  for i in range(num_batches): 
    #embed,logits = net(graph.to(device),input_features.to(device),batches[i].to(device)) #in case of seq gnn model
    embed,logits = net(input_features.to(device),ax.to(device),a2x.to(device),batches[i].to(device)) #logits -- [36]
    loss = BCEloss(logits,batch_labels[i].type(torch.FloatTensor).to(device))
    optimizer.zero_grad()
    loss.backward()
    optimizer.step()
    batch_loss.append(loss.item())
    sg = sig(logits)
    predicted_labels = [1 if sg[j]>0.5 else 0 for j in range(len(logits))]
    acc = sum([1 if predicted_labels[j] == batch_labels[i,j] else 0 for j in range(batch_size)])/batch_size
    training_acc.append(acc)
    print ("epoch = %d | batch = %d | batch_loss = %f | Batch_Acc = %f | "%(epoch,i,loss.item(),acc),end=" ")
    TP = sum([1 if ((predicted_labels[j] == 1) and (batch_labels[i,j]==1)) else 0 for j in range(batch_size)])
    FP = sum([1 if ((predicted_labels[j] == 1) and (batch_labels[i,j]==0)) else 0 for j in range(batch_size)])
    FN = sum([1 if ((predicted_labels[j] == 0) and (batch_labels[i,j]==1)) else 0 for j in range(batch_size)])
    TN = sum([1 if ((predicted_labels[j] == 0) and (batch_labels[i,j]==0)) else 0 for j in range(batch_size)])
    print (TP,FP,FN,TN,end=" ")
    try:
      pre = TP/(TP+FP)
    except:
      pre = 0.95
    rec = TP/(TP+FN)
    try:
      f1 = 2*(pre*rec)/(pre+rec)
    except:
      f1 = 0
    F1_scores.append(f1)
    print ("| Precision = %f | Recall = %f | F1 = %f"%(pre,rec,f1))
  epoch_loss.append(batch_loss)
  avg_epoch_loss.append(sum(batch_loss)/num_batches)
  print(" ------- Epoch {:05d} | Avg_BatchLoss {:.4f} | Avg_test_loss = {:.4f}".format(epoch, sum(batch_loss)/num_batches,0))

index = 0
#plt.plot(range(len(training_acc)-index),training_acc[index:])
plt.plot(range(len(F1_scores)-index),F1_scores[index:],color = "crimson")
plt.plot(range(len(training_acc)),training_acc)

#save_variable
#define your model directory
save_variable(F1_scores,model_directory+"F1_scores.p")
save_variable(training_acc,model_directory+"training_acc.p")

#----------Evaluation--------------------------
#testing_acc_BCE
net.eval()
embed,logits = net(ax.to(device),a2x.to(device),input_features.to(device),testing_links.to(device)) #for sign
#_,logits = net(graph.to(device),input_features.to(device),testing_links.to(device)) #for seq gnn
sg = sig(logits)
predicted_labels = [1 if sg[j]>0.5 else 0 for j in range(len(logits))]
testing_acc = sum([1 if predicted_labels[j] == testing_labels[j] else 0 for j in range(len(testing_links))])/len(testing_links)
TP = sum([1 if ((predicted_labels[j] == 1) and (testing_labels[j]==1)) else 0 for j in range(len(testing_links))])
FP = sum([1 if ((predicted_labels[j] == 1) and (testing_labels[j]==0)) else 0 for j in range(len(testing_links))])
FN = sum([1 if ((predicted_labels[j] == 0) and (testing_labels[j]==1)) else 0 for j in range(len(testing_links))])
TN = sum([1 if ((predicted_labels[j] == 0) and (testing_labels[j]==0)) else 0 for j in range(len(testing_links))])
print (testing_acc,TP,FP,FN,TN,end=" | ")
pre = TP/(TP+FP)
rec = TP/(TP+FN)  
print ("Precision = %f | Recall = %f | F1 = %f"%(pre,rec,2*(pre*rec)/(pre+rec)))

fpr, tpr, thresholds = sklearn.metrics.roc_curve(y_true = testing_labels, y_score = sg.detach().to("cpu"), pos_label = 1) #positive class is 1; negative class is 0
auroc = sklearn.metrics.auc(fpr, tpr)
print (auroc)
#fpr_sage,tpr_sage,auroc_sage = fpr,tpr,auroc

fig1 = plt.figure(figsize=(4, 3))
c = fig1.add_subplot(1, 1, 1)
sns.set_style(style='whitegrid')
sns.lineplot(x = fpr,y = tpr)
plt.plot([0, 1], [0, 1],'-.')
c.legend(['SIGN AUROC = 0.924','Random classifier = 0.5'])
c.set(xlabel='False positive rate', ylabel='True positive rate',title='ROC Curves')
fig1.savefig("comparison.png",dpi=500)
fig1.savefig("svg_comparison.svg",dpi = 500)

save_model(net,"SIGN_Models/Model_9/model_decagon_512_2lac_4060_tanh_1")
save_variable(embed.detach().to('cpu'),model_directory+"embed.p")

#----------------loading the pre trained model------------- 
# The detailed discussion on how to use our model to predict the drugs for a disease is given in other file
model_directory = 'SIGN_Models/Model_7/'
graph = load_variable(model_directory+"Data/graph.p")
input_features = load_variable(model_directory+"Data/input_features.p") #give in a path of input_features
nodes_mapping = load_variable(model_directory+"Data/nodes_mapping.p") #path for nodes_mapping
#we can get the ax and a2x from inp feat -- load A_tilda
model_directory = 'SIGN_Models/Model_9/'
empty_model = SIGN()
net = load_model_on_cpu(empty_model,model_directory+"model_decagon_512_2lac_4060_tanh").to(device)
X_embedded = load_variable(model_directory+"embed.p")

#loading a seq gnn model
model_directory = 'SIGN_Models/Model_9/'
input_features = load_variable(model_directory+"Data/input_features.p")
empty_model = Seq_GNN()
net = load_model_on_cpu(empty_model,model_directory+"model_GAT").to(device)

def get_DRKG_scores_dict(relation_embedding,dis_list,input_fetaures): #list of diseases - in Disease::MESH:D###### format
  dis_batches,drug_dict = get_disease_batches(nodes_mapping,dis_list)
  dis_batches = torch.LongTensor(dis_batches)
  dictionaries = []
  for i in range(len(dis_batches)):
    scores = []
    for link in dis_batches[i]:
      d = 12-np.linalg.norm(input_features[link[0]]+relation_embedding-input_features[link[1]])
      score = np.log(1/(1+np.exp(-d)))
      scores.append(score)
    dct = dict()
    for j in range(len(scores)):
      x = get_node_name(dis_batches[i,j,0].item())
      if (x in drug_dict.keys()) : 
        dct[drug_dict[x][0]] = scores[j]
      else : 
        dct[x] = scores[j]
    
    dictionaries.append(dct)
  return dictionaries

def predict(dis_list): #list of diseases - in Disease::MESH:D###### format
  dis_batches,drug_dict = get_disease_batches(nodes_mapping,dis_list)
  dis_batches = torch.LongTensor(dis_batches)
  dictionaries_norm = []
  for i in range(len(dis_batches)):
    embed,logits = net(input_features.to(device),ax.to(device),a2x.to(device),dis_batches[i].to(device))
    #_,logits = net(graph.to(device),input_features.to(device),dis_batches[i].to(device))
    probs = standardize(logits)
    dct_norm = dict ()
    for j in range(len(probs)):
      x = get_node_name(dis_batches[i,j,0].item())
      if (x in drug_dict.keys()) : 
        dct_norm[drug_dict[x][0]] = probs[j].item()
      else : 
        dct_norm[x] = probs[j].item()
    print (i)
    dictionaries_norm.append(dct_norm)
  return embed,dictionaries_norm

def get_zscore_dict(dis_list):
  dis_batches,drug_dict = get_disease_batches(nodes_mapping,dis_list)
  dis_batches = torch.LongTensor(dis_batches)
  dictionaries_z = []
  for i in range(len(dis_batches)):
    logits_z = []
    for j in range(len(dis_batches[i])):
      a = get_Zscore(all_genes,dis_batches[i,j,1].item(),dis_batches[i,j,0].item(),degreewise_gene_dct,gene_network,num_seeds=600)
      logits_z.append(a)
      print(j)
    dct_z = dict ()
    for j in range(len(logits_z)):
      x = get_node_name(dis_batches[i,j,0].item())
      if (x in drug_dict.keys()) : 
        dct_z[drug_dict[x][0]] = logits_z[j]
      else : 
        dct_z[x] = logits_z[j]
    dictionaries_z.append(dct_z)
  return dictionaries_z

def atc_validation(predicted_list):
  first_level_atc = []
  for drug in predicted_list:
    for atc in atc_dictionary[drug]:
      first_level_atc.append(atc[0])
  return first_level_atc

def get_top_k_names(dct,k=50):
  lst = sorted(dct.items(),key=lambda t:t[1])[:k]
  names = [tup[0] for tup in lst]
  return names

def get_node_id_from_name(name): #node id in the graph - from it's drug name
  for key in drug_name_dict.keys():
    if (drug_name_dict[key][0]==name):
      id = get_node_id(key)
      break
  return id

def get_rank(dct,key,probs=True): #probs=False meaning Ranks are directly proportional to the scores
  lst = sorted(dct.items(),key=lambda t:t[1])[::-1] if (probs) else sorted(dct.items(),key=lambda t:t[1])
  for i in range(len(lst)):
    if (key==lst[i][0]):
      break
  return i+1

def get_group_from_compound(drug_name): #give in compound in the form of name = 'Dexamethasone' form 
  group = ''
  for key,value in drug_name_dict.items():
    if (value[0]==drug_name):
      group = value[1]
      break
  return group
  
model_directory = 'SIGN_Models/Model_7/'
training_links1 = load_variable(model_directory+"Data/training_links.p")

training_dis = set()
for link in training_links1:
  training_dis.add(link[1])
print (len(training_dis))
training_dis = list(training_dis)

#getting ranks for of the drugs on the testing diseases-----------------
testing_diseases = [12419,40093,24354,34794,590,9611,20473,13324,24153,2008,10292,23186,37549,35522,25773,9934,12741,33207,40770]
testing_diseases = [get_node_name(i) for i in testing_diseases]
print (len(testing_diseases))
testing_drugs = ["Acyclovir","Fluorouracil","Prednisone","Nitroglycerin","Bleomycin","Atropine","Liothyronine","Lovastatin","Colchicine","Tetracosactide","Hyoscyamine","Albumin human","Niacin","Hyoscyamine","Botulinum toxin type A","Cefuroxime","Ceftolozane","Azacitidine","Alemtuzumab"]

dictionaries_drkg = get_DRKG_scores_dict(db_treats_relation_embedding,[get_node_name(32809)],input_features)
#embed,dictionaries_sign = predict(testing_diseases)
#embed,dictionaries_sign = predict([get_node_name(14347)])
#embed,dictionaries_sign_covid = predict(covid_disease_list)
Z_rectal = get_zscore_dict([get_node_name(40093)])
save_variable(Z_rectal,model_directory+"Z_rectal.p")
#save_variable(dictionaries_sign_covid_1,"SIGN_Models/Model_9/dictionaries_sign_covid_1.p")

x = dictionaries_sign_covid[10]
sorted(get_ranks_covid_drugs(x,probs=True).items(),key=lambda t:t[1][0])

x = dictionaries_sign_covid[0]
#x = Z_leukemia[0]
#x = dictionaries_drkg[0]
sorted(x.items(),key=lambda t:t[1])[::-1][:15]
#Counter(atc_validation(get_top_k_names(x)))

for i in range(len(testing_drugs)):
  print (testing_drugs[i],get_rank(dictionaries_sign[i],testing_drugs[i]),sep = "--")
else:
  print ("Cefotaxime",get_rank(dictionaries_sign[-3],"Cefotaxime"),sep = "--")
  print ("Decitabine",get_rank(dictionaries_sign[-2],"Decitabine"),sep = "--")

#Performance (ranks) check on the trained diseases as well	
wb = openpyxl.Workbook()
wb.save("Model_Performance_1.xlsx")
for j in range(0,75):
  dis_id = training_dis[j]
  #print (training_links[id][1],get_node_name(training_links[id][1]))
  #print (training_links[id][0],get_node_name(training_links[id][0]),drug_name_dict[get_node_name(training_links[id][0])][0])
  d = []
  for edge in graph_nx.edges(dis_id):
    name = get_node_name(edge[1])
    if (re.search(r"Compound",name,re.I)):
      #print (drug_name_dict[name][0])
      d.append(drug_name_dict[name][0])
  _,dictionaries_sign = predict([get_node_name(dis_id)]) #give in a disease list
  wb = openpyxl.load_workbook("Model_Performance_1.xlsx")
  sheet = wb.active
  r_max = sheet.max_row
  sheet.cell(row=r_max+1,column=1).value = j+1
  sheet.cell(row=r_max+1,column=2).value = dis_id
  sheet.cell(row=r_max+1,column=3).value = get_node_name(dis_id)
  for i in range(len(d)):
    sheet.cell(row=r_max+1+i,column=1).value = j+1
    sheet.cell(row=r_max+1+i,column=4).value = d[i]
    sheet.cell(row=r_max+1+i,column=5).value = get_rank(dictionaries_sign[0],d[i],probs = True)
    #sheet.cell(row=r_max+1+i,column=6).value = get_rank(dictionaries_drkg[0],d[i],probs = True)
  print ("j = %d done..."%(j))
  wb.save("Model_Performance_1.xlsx")

excel_covid_clinical_results(dictionaries_sign_covid,True,"Covid_predictions_clinical_full_n_final_final.xlsx")
excel_results(dictionaries_sign_covid,"Covid_predictions_full_n_final_final.xlsx")

#-------------------------------------Visualization------------------------------------#heatmap

def sort_based_on_phase_trials(drug_list):
  wb = openpyxl.load_workbook("Covid_clinical_drugs.xlsx").active
  dct_complete = dict()
  for i in range(1,wb.max_row+1):
    dct_complete[wb.cell(row=i,column=1).value] = wb.cell(row=i,column=2).value 
  drug_phase_pair = []
  for drug in drug_list:
    if (dct_complete[drug]==None):print (drug)
    drug_phase_pair.append((drug,dct_complete[drug]))
  sorted_drug_list = []
  for item in sorted(drug_phase_pair,key=lambda t:t[1])[::-1]:
    sorted_drug_list.append(item[0])
  return sorted_drug_list

def group_based_on_atc(drug_list):
  drug_atc_pair = []
  for drug in drug_list:
    if(len(atc_dictionary[drug])>0):
      drug_atc_pair.append((drug,atc_dictionary[drug][0][0]))
    else:
      drug_atc_pair.append((drug,'Z'))
  sorted_drug_list = []
  for item in sorted(drug_atc_pair,key=lambda t:t[1]):
    sorted_drug_list.append(item[0])
  return sorted_drug_list

def get_heatmap_data(dictionaries,k=10,clinical=True,probs=True): #get the top k
    top_k_clinical_drugs = set()
    top_k_drugs = set ()
    for dct in dictionaries:
      for drug,x in sorted(get_ranks_covid_drugs(dct,probs).items(),key=lambda t:t[1])[:k]:
        if (x[0]<k) : top_k_clinical_drugs.add(drug)
      for drug,score in sorted(dct.items(),key=lambda t:t[1])[::-1][:k]:
        top_k_drugs.add(drug)
    drugs_list = list(top_k_clinical_drugs) if (clinical) else list(top_k_drugs)
    drugs_list = drugs_list if (clinical) else group_based_on_atc(drugs_list) 
    print ("Total drugs -- ",end=" ")
    print (len(drugs_list))
    #print (*list(drugs_list),sep='\n',end='\n')
    wb_new = openpyxl.Workbook()
    sheet = wb_new.active
    for i in range(1,len(drugs_list)+1):
      sheet.cell(row=i,column=1).value = drugs_list[i-1]
      for j in range(2,len(dictionaries)+2):
        r = get_rank(dictionaries[j-2],drugs_list[i-1])
        sheet.cell(row=i,column=j).value = r
    wb_new.save("/content/drive/My Drive/Work_Drug_Repurpose/Results.xlsx")
    data = np.empty([sheet.max_row,len(dictionaries)])
    for i in range(1,sheet.max_row+1):
      for j in range(2,len(dictionaries)+2):
        data[i-1,j-2] = sheet.cell(row=i,column=j).value
    return data,drugs_list    

def star_clinical_drugs(drug_list):
  wb = openpyxl.load_workbook("Covid_clinical_drugs.xlsx").active
  all_clinical_drugs = [wb.cell(row=i,column=1).value for i in range(1,wb.max_row+1)]
  for i in range(len(drug_list)):
    if (drug_list[i] in all_clinical_drugs) : 
      drug_list[i] = "**"+drug_list[i]
  return drug_list



data,drugs_sign = get_heatmap_data(dictionaries_sign_covid,k=10,clinical=False,probs=True)
data,drugs_drkg = get_heatmap_data(dictionaries_drkg,k=50,clinical=True,probs=True)
drugs_starred = star_clinical_drugs(drugs_sign)

Counter(atc_validation(drugs_drkg))
Counter(atc_validation(drugs_sign))

drugs_starred[drugs_starred.index("Varicella Zoster Vaccine (Live/attenuated)")] = "Varicella Zoster Vaccine"
drugs_starred[drugs_starred.index("Ebola Zaire vaccine (live, attenuated)")] = "Ebola Zaire vaccine"
drugs_starred[drugs_starred.index("Collagenase clostridium histolyticum")] = "Collagenase clostridium"
drugs_starred[drugs_starred.index("Samarium (153Sm) lexidronam")] = "Samarium lexidronam"
drugs_starred[drugs_starred.index("Human Rho(D) immune globulin")] = "Rho(D) immune globulin"
#fig.savefig()

disease_labels = ['SARS-CoV2-E','SARS-CoV2-M','SARS-CoV2-N','SARS-CoV2-spike','nsp1','nsp10','nsp11','nsp12','nsp13','nsp14','nsp15','nsp2','nsp4','nsp5','nsp5_C145A','nsp6','nsp7','nsp8','nsp9','orf10','orf3a','orf3b','orf6','orf7a','orf8','orf9b','orf9c','SARS','IBV','MERS','CoV-229E','CoV-NL63','MHV']

#Generate Heat map
fig,ax = plt.subplots(figsize=(10,16))
sns.set(font_scale=0.5)
ax = sns.heatmap(data,xticklabels=disease_labels,yticklabels=drugs_starred,linewidth=0.25,cmap="Reds_r",cbar=True,cbar_kws = dict(use_gridspec=False,location="right",shrink=0.85))
ax.tick_params(labelright = False,labelleft = True,labeltop=False,rotation=0)
plt.xticks(rotation=90)
for i, ticklabel in enumerate(ax.yaxis.get_majorticklabels()):
  if (i<drugs_sign.index('**Cholecalciferol')+1) : ticklabel.set_color('forestgreen')
  elif (i<drugs_sign.index('**Acetylsalicylic acid')+1) : ticklabel.set_color('deeppink')
  elif (i<drugs_sign.index('**Rosuvastatin')+1) : ticklabel.set_color('navy')
  elif (i<drugs_sign.index('Hydroquinone')+1) : ticklabel.set_color('crimson')
  elif (i<drugs_sign.index('Levonorgestrel')+1) : ticklabel.set_color('lightseagreen')
  elif (i<drugs_sign.index('Rimexolone')+1) : ticklabel.set_color('rosybrown')
  elif (i<drugs_sign.index('**Rubella virus vaccine')+1) : ticklabel.set_color('royalblue')
  elif (i<drugs_sign.index('Mechlorethamine')+1) : ticklabel.set_color('darkorange')
  elif (i<drugs_sign.index('Ataluren')+1) : ticklabel.set_color('darkorchid')
  elif (i<drugs_sign.index('Carbamazepine')+1) : ticklabel.set_color('maroon')
  elif (i<drugs_sign.index('**Ivermectin')+1) : ticklabel.set_color('teal')
  elif (i<drugs_sign.index('**Prednisolone')+1) : ticklabel.set_color('orangered')
  elif (i<drugs_sign.index('Ranibizumab')+1) : ticklabel.set_color('darkgreen')
  elif (i<drugs_sign.index('Samarium lexidronam')+1) : ticklabel.set_color('purple')
  else : ticklabel.set_color('dimgray')
plt.show()

#fig.savefig("Images/heatmap_900.png",dpi=900,transparent=True)
#fig.savefig("Images/svg_heatmap_500.svg",dpi=500,transparent=True)
#fig.savefig("Images/eps_heatmap_500.eps",dpi=500,transparent=True)
#fig.savefig("Images/heatmap_500.png",dpi=500,transparent=True)
fig.savefig("962.svg",dpi=962,transparent=True)

def drug_gene_dis_linkages(dis_id,drug_id):
  C = get_associated_genes(all_genes,dis_id)
  T = get_associated_genes(all_genes,drug_id)
  wb = openpyxl.Workbook()
  sheet=wb.active
  sheet.cell(row=1,column=1).value = "Source"
  sheet.cell(row=1,column=2).value = "Target"
  edges = []
  i=2
  for c in C.union(T):
    for edge in list(graph_nx.edges(nbunch=c)):
      if (re.search(r"Gene",get_node_name(edge[1]),re.I)):
        sheet.cell(row=i,column=1).value = get_node_name(edge[0])  
        sheet.cell(row=i,column=2).value = get_node_name(edge[1])
        i+=1
  for c in C:
    sheet.cell(row=i,column=1).value = get_node_name(dis_id)  
    sheet.cell(row=i,column=2).value = get_node_name(c)
    i+=1
  for t in T:
    sheet.cell(row=i,column=1).value = get_node_name(drug_id)  
    sheet.cell(row=i,column=2).value = get_node_name(t)
    i+=1
  wb.save("/content/drive/My Drive/Work_Drug_Repurpose/SIGN_Models/Post_Analysis/IBV_Tacrolimus_dgdl.xlsx")

#change the file name in the function everytime
drug_gene_dis_linkages(get_node_id('Disease::MESH:D001351'),get_node_id("Compound::DB00864"))

def drug_gene_anat_gene_dis_linkages(dis_id,drug_id):
  C = get_associated_genes(all_genes,dis_id)
  T = get_associated_genes(all_genes,drug_id)
  wb = openpyxl.Workbook()
  sheet=wb.active
  sheet.cell(row=1,column=1).value = "Source"
  sheet.cell(row=1,column=2).value = "Target"
  edges = []
  i=2
  for c in C.union(T):
    for edge in list(graph_nx.edges(nbunch=c)):
      if (re.search(r"Anatomy",get_node_name(edge[1]),re.I)):
        sheet.cell(row=i,column=1).value = get_node_name(edge[0])  
        sheet.cell(row=i,column=2).value = get_node_name(edge[1])
        i+=1
  for c in C:
    sheet.cell(row=i,column=1).value = get_node_name(dis_id)  
    sheet.cell(row=i,column=2).value = get_node_name(c)
    i+=1
  for t in T:
    sheet.cell(row=i,column=1).value = get_node_name(drug_id)  
    sheet.cell(row=i,column=2).value = get_node_name(t)
    i+=1
  wb.save("/content/drive/My Drive/Work_Drug_Repurpose/SIGN_Models/Post_Analysis/Cov229_favirpiravir_dgagdl.xlsx")

#change the file name in the function everytime
drug_gene_anat_gene_dis_linkages(get_node_id('Disease::MESH:D028941'),get_node_id("Compound::DB12466"))
